import os
import sys
import subprocess
import re
import glob

# === БЛОК АВТОНОМНОСТИ: Установка библиотек ===
def ensure_dependencies():
    required = ["pdfplumber", "openpyxl", "pandas"]
    for pkg in required:
        try:
            __import__(pkg if pkg != "pdfplumber" else "pdfplumber")
        except ImportError:
            print(f"Установка компонента: {pkg}...")
            try:
                subprocess.check_call([sys.executable, "-m", "pip", "install", pkg])
            except Exception as e:
                print(f"Ошибка установки {pkg}: {e}")

ensure_dependencies()

import pdfplumber
import pandas as pd
import openpyxl
from openpyxl.styles import Font

# === ЛОГИКА ПАРСИНГА ===

def normalize_text(text):
    """Очистка текста для надежного сравнения."""
    if not text: return ""
    return re.sub(r"\s+", " ", str(text)).strip()

def extract_addendum_mapping(pdf_path, models_from_xlsx):
    """Связывает модели из XLSX с номерами товаров из Дополнения."""
    mapping = {}
    print(f"Анализ дополнения: {os.path.basename(pdf_path)}")
    
    with pdfplumber.open(pdf_path) as pdf:
        full_text = ""
        for page in pdf.pages:
            full_text += page.extract_text() + "\n"
        
        # Разбиваем на блоки по товарам "------------ Товар № X ------------"
        blocks = re.split(r"-+\s*Товар\s*№\s*(\d+)\s*-+", full_text)
        
        # blocks[0] - текст до первого товара
        # blocks[1], blocks[2]... - пары (номер, текст блока)
        for i in range(1, len(blocks), 2):
            item_num = blocks[i]
            block_text = blocks[i+1]
            
            # Ищем в тексте блока любую из моделей, которые мы нашли в Excel
            for model in models_from_xlsx:
                # Ищем модель как отдельное слово или подстроку
                if model.upper() in block_text.upper():
                    mapping[model] = item_num
                    print(f"  [Доп] Модель {model} -> Товар № {item_num}")
                    
    return mapping

def extract_gtd_mapping(pdf_path):
    """Извлекает маппинг {Номер_товара: Код_ТНВЭД} из основной ГТД."""
    mapping = {}
    print(f"Анализ ГТД: {os.path.basename(pdf_path)}")
    
    with pdfplumber.open(pdf_path) as pdf:
        full_text = ""
        for page in pdf.pages:
            full_text += page.extract_text() + "\n"
        
        # В ГТД ячейки 32 и 33 часто содержат значения в конце описания товара (графа 31).
        # Ищем паттерны: цифра (номер товара) и длинное число (код ТНВЭД 10 знаков)
        # Пример из дампа: "описание 1-... : 3 8479810000 С N"
        
        # Регулярка ищет номер товара (1-2 цифры) и следующий за ним 10-значный код
        # Мы ищем их в контексте слова "описание" или просто по всему тексту
        matches = re.findall(r":\s*(\d{1,2})\s+(\d{10})", full_text)
        
        if not matches:
            # Fallback: ищем просто группы цифр в конце строк описания
            matches = re.findall(r"\s+(\d{1,2})\s+(\d{10})\s+[A-Z]", full_text)

        for num, code in matches:
            mapping[num] = code
            print(f"  [ГТД] Товар № {num} -> Код ТНВЭД {code}")
            
    return mapping

# === ЛОГИКА ПОИСКА ФАЙЛОВ ===

def find_files_in_folder(folder):
    """Ищет тройку файлов в конкретной папке."""
    xlsx = glob.glob(os.path.join(folder, "Табл*.xlsx"))
    gtd = glob.glob(os.path.join(folder, "ГТД*.pdf"))
    dop = glob.glob(os.path.join(folder, "Доп*.pdf"))
    return xlsx, gtd, dop

def scan_all_folders(root_dir, excel_models):
    """Рекурсивно собирает маппинг {Модель: Код} изо всех подпапок."""
    global_model_to_tnved = {}
    
    print(f"Запуск глубокого сканирования в: {root_dir}")
    
    # Сначала проверяем саму корневую папку
    for current_dir, dirs, files in os.walk(root_dir):
        _, gtd_list, dop_list = find_files_in_folder(current_dir)
        
        if gtd_list and dop_list:
            print(f"\n Найдена пара документов в папке: {os.path.relpath(current_dir, root_dir)}")
            try:
                # Парсим текущую папку
                m_to_n = extract_addendum_mapping(dop_list[0], excel_models)
                n_to_t = extract_gtd_mapping(gtd_list[0])
                
                # Соединяем в локальный маппинг для этой папки
                for m, n in m_to_n.items():
                    if n in n_to_t:
                        global_model_to_tnved[m] = {
                            "num": n,
                            "code": n_to_t[n],
                            "path": os.path.abspath(dop_list[0]) # Ссылка на ДОПОЛНЕНИЕ
                        }
            except Exception as e:
                print(f"  [!] Ошибка при обработке папки {current_dir}: {e}")
                
    return global_model_to_tnved

def main():
    print("\n" + "="*40)
    print("  ПАРСЕР ТН ВЭД (Универсальный: Папка или Сканер)")
    print("="*40 + "\n")
    
    root_dir = os.getcwd()
    xlsx_files, gtd_files, dop_files = find_files_in_folder(root_dir)
    
    # 1. Поиск основной таблицы
    if not xlsx_files:
        print("ОШИБКА: Не найдена таблица Excel (Табл*.xlsx) в текущей папке!")
        input("\nНажмите Enter...")
        return
    
    table_path = xlsx_files[0]
    print(f"Загрузка таблицы: {os.path.basename(table_path)}")
    wb = openpyxl.load_workbook(table_path)
    ws = wb.active
    
    # 2. Сбор моделей из Excel для поиска
    excel_models = set()
    for row in range(2, ws.max_row + 1):
        val = ws.cell(row=row, column=2).value
        if val:
            m = str(val).strip().split(' ')[0]
            if len(m) > 2: excel_models.add(m)
    
    print(f"Моделей для поиска в Excel: {len(excel_models)}")

    # 3. Сбор данных (Режим 1: Все в одной папке)
    global_mapping = {}
    if gtd_files and dop_files:
        print("Режим: Все файлы в одной папке.")
        m_to_n = extract_addendum_mapping(dop_files[0], excel_models)
        n_to_t = extract_gtd_mapping(gtd_files[0])
        for m, n in m_to_n.items():
            if n in n_to_t:
                global_mapping[m] = {
                    "num": n,
                    "code": n_to_t[n],
                    "path": os.path.abspath(dop_files[0]) # Ссылка на ДОПОЛНЕНИЕ
                }
    else:
        # Режим 2: Сканирование подпапок
        print("Режим: Сканирование подпапок...")
        global_mapping = scan_all_folders(root_dir, excel_models)

    if not global_mapping:
        print("\nОШИБКА: Не удалось собрать данные ни из одной папки!")
        print("Убедитесь, что рядом с ГТД лежит Дополнение к ней.")
        input("\nНажмите Enter...")
        return

    # 4. Обновление таблицы
    # Динамический поиск колонок
    col_map = {}
    for row in range(1, min(10, ws.max_row + 1)):
        for col in range(1, ws.max_column + 1):
            val = str(ws.cell(row=row, column=col).value).strip().lower()
            if "наименование" in val or "модель" in val: col_map["name"] = col
            if "№ товара" in val or "номер товара" in val: col_map["num"] = col
            if "тн вэд" in val or "код товара" in val or "тнвэд" in val: col_map["code"] = col
    
    if not all(k in col_map for k in ["name", "num", "code"]):
        print(f"ОШИБКА: Не удалось найти все колонки в Excel! Найдено: {col_map}")
        input("\nНажмите Enter...")
        return

    updated_count = 0
    print(f"\nИтого найдено уникальных кодов для моделей: {len(global_mapping)}")
    print(f"Колонки: Имя={col_map['name']}, Номер={col_map['num']}, Код={col_map['code']}")
    print("Обновление Excel...")
    
    for row in range(1, ws.max_row + 1):
        cell_val = ws.cell(row=row, column=col_map["name"]).value
        if not cell_val: continue
        
        # Пропускаем строку, если это заголовок
        clean_val = str(cell_val).strip().lower()
        if clean_val in ["наименование", "модель", "товар"]:
            continue
            
        m_name = str(cell_val).strip().split(' ')[0]
        if m_name in global_mapping:
            data = global_mapping[m_name]
            
            # Номер товара + Ссылка (Дополнение)
            cell_num = ws.cell(row=row, column=col_map["num"])
            if str(cell_num.value).strip().lower() not in ["№ товара", "номер", "№"]:
                cell_num.value = data["num"]
                cell_num.hyperlink = data["path"]
                cell_num.font = Font(color="0000FF", underline="single")
            
            # Код ТН ВЭД
            cell_tnved = ws.cell(row=row, column=col_map["code"])
            if str(cell_tnved.value).strip().lower() not in ["тн вэд", "тнвэд", "код"]:
                cell_tnved.value = data["code"]
            
            updated_count += 1
            print(f"  OK: {m_name} -> №{data['num']} (код {data['code']})")

    output_path = table_path.replace(".xlsx", "_готово.xlsx")
    wb.save(output_path)
    
    print("\n" + "-"*40)
    print(f"УСПЕХ! Всего обновлено строк: {updated_count}")
    print(f"Результат: {os.path.basename(output_path)}")
    print("-"*40)
    input("\nНажмите Enter...")

if __name__ == "__main__":
    main()
