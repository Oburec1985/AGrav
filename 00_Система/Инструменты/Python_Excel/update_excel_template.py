import argparse
import csv
import ctypes
import os
import re
import subprocess
import sys

def ensure_dependencies():
    """Checks and installs missing python packages automatically."""
    required = ["openpyxl"]
    for pkg in required:
        try:
            __import__(pkg)
        except ImportError:
            print(f"Installing missing dependency: {pkg}...")
            try:
                subprocess.check_call([sys.executable, "-m", "pip", "install", pkg])
            except Exception as e:
                print(f"Error installing {pkg}: {e}")

ensure_dependencies()

import openpyxl


def column_to_index(value):
    """Converts Excel column letter or 1-based number to a column index."""
    if isinstance(value, int):
        return value

    text = str(value).strip()
    if text.isdigit():
        return int(text)

    result = 0
    for char in text.upper():
        if not ("A" <= char <= "Z"):
            raise ValueError(f"Invalid column: {value}")
        result = result * 26 + (ord(char) - ord("A") + 1)
    return result


def normalize_name(name):
    """Normalizes names for robust matching between Excel and result.csv."""
    text = (name or "").upper().replace("Ё", "Е")
    text = re.sub(r"[^А-ЯA-Z0-9\s\.-]", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def find_default_csv():
    """Finds parser output with support for both result.csv and results.csv."""
    for filename in ("result.csv", "results.csv"):
        if os.path.exists(filename):
            return filename
    return "result.csv"


def expand_windows_short_path(path):
    """Expands an 8.3 Windows short path back to its long form when possible."""
    if os.name != "nt" or not path:
        return path

    try:
        get_long_path_name = ctypes.windll.kernel32.GetLongPathNameW
        buffer = ctypes.create_unicode_buffer(32768)
        result = get_long_path_name(str(path), buffer, len(buffer))
        if result:
            return buffer.value
    except Exception:
        pass

    return path


def resolve_excel_file(excel_ref, search_dir="."):
    """Resolves an Excel file by exact path or by case-insensitive filename substring."""
    excel_ref = expand_windows_short_path(excel_ref)
    if os.path.exists(excel_ref):
        return excel_ref

    needle = str(excel_ref).strip().lower()
    if not needle:
        raise ValueError("Excel file substring is empty")

    matches = []
    for filename in os.listdir(search_dir):
        if filename.startswith("~$"):
            continue
        if not filename.lower().endswith(".xlsx"):
            continue
        if filename.lower().endswith("_updated.xlsx"):
            continue
        if needle in filename.lower():
            matches.append(os.path.join(search_dir, filename))

    matches.sort(key=lambda path: os.path.basename(path).lower())
    if not matches:
        raise FileNotFoundError(
            f"Excel file not found by path or substring: {excel_ref}. Search dir: {os.path.abspath(search_dir)}"
        )

    if len(matches) > 1:
        print("Found multiple Excel files, using the first one:")
        for path in matches:
            print(f"  - {path}")

    return matches[0]


def load_mapping_from_csv(csv_path):
    """Loads {employee_name: amount} from parser CSV output."""
    if not os.path.exists(csv_path):
        raise FileNotFoundError(f"CSV file not found: {csv_path}")

    mapping = {}
    duplicates = []

    with open(csv_path, "r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f, delimiter=";")
        if not reader.fieldnames or "ФИО" not in reader.fieldnames or "Значение" not in reader.fieldnames:
            raise ValueError("CSV must contain columns: ФИО;Значение")

        for row in reader:
            name = (row.get("ФИО") or "").strip()
            value_text = (row.get("Значение") or "").strip().replace(",", ".")
            if not name or not value_text:
                continue

            try:
                value = float(value_text)
            except ValueError:
                print(f"WARNING: skipped invalid amount for {name}: {value_text}")
                continue

            if name in mapping:
                duplicates.append(name)
            mapping[name] = value

    if duplicates:
        print(f"WARNING: duplicate names in CSV, last value was used: {sorted(set(duplicates))}")

    return mapping


def match_mapping_key(excel_name, mapping):
    """Finds a mapping key for an Excel name."""
    normalized_excel = normalize_name(excel_name)
    normalized_mapping = {normalize_name(key): key for key in mapping}

    if normalized_excel in normalized_mapping:
        return normalized_mapping[normalized_excel]

    for normalized_key, original_key in normalized_mapping.items():
        if len(normalized_excel) > 3 and (
            normalized_excel.startswith(normalized_key) or normalized_key.startswith(normalized_excel)
        ):
            return original_key

    return None


def update_excel_with_mapping(
    file_path,
    sheet_name,
    name_col,
    target_col,
    mapping,
    start_row=1,
    output_path=None,
    value_divisor=1.0,
):
    """
    Updates an Excel file using a {name: value} mapping.

    Args:
        file_path: Path to source .xlsx file.
        sheet_name: Worksheet name.
        name_col: 1-based column index with employee names.
        target_col: 1-based column index to update.
        mapping: Dictionary {employee name: amount}.
        start_row: First row to scan.
        output_path: Optional output file path. Defaults to *_updated.xlsx.
        value_divisor: Divide values before writing. Default: 1.0.
    """
    file_path = expand_windows_short_path(file_path)
    if output_path:
        output_path = expand_windows_short_path(output_path)

    if not os.path.exists(file_path):
        raise FileNotFoundError(f"Excel file not found: {file_path}")

    wb = openpyxl.load_workbook(file_path)
    if sheet_name is None:
        sheet = wb.active
        print(f"Using active sheet: {sheet.title}")
    elif sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet {sheet_name} not found. Available: {wb.sheetnames}")
    else:
        sheet = wb[sheet_name]
    updated_count = 0
    missed = []

    for row in range(start_row, sheet.max_row + 1):
        name_cell = sheet.cell(row=row, column=name_col).value
        if not name_cell:
            continue

        name_str = str(name_cell).strip()
        mapping_key = match_mapping_key(name_str, mapping)
        if not mapping_key:
            missed.append(name_str)
            continue

        value = mapping[mapping_key] / value_divisor
        sheet.cell(row=row, column=target_col).value = value
        updated_count += 1
        print(f"Updated row {row}: {name_str} <- {mapping_key} = {value}")

    if output_path is None:
        root, ext = os.path.splitext(file_path)
        output_path = f"{root}_updated{ext}"

    wb.save(output_path)
    print(f"Done. Updated {updated_count} rows. Saved as {output_path}")

    if missed:
        print(f"WARNING: names not found in CSV mapping ({len(missed)}):")
        for name in missed:
            print(f"  - {name}")

    return output_path


def build_arg_parser():
    parser = argparse.ArgumentParser(
        description="Update an Excel template using salary values from result.csv/results.csv."
    )
    parser.add_argument("excel_file", nargs="?", default=None, help="Path to .xlsx file or filename substring. If missing, searches for .xlsx nearby.")
    parser.add_argument("--sheet", default=None, help="Worksheet name. Default: active sheet.")
    parser.add_argument("--name-col", default="B", help="Employee name column, e.g. B or 2. Default: B.")
    parser.add_argument("--target-col", default="H", help="Column to write values, e.g. H or 8. Default: H.")
    parser.add_argument("--start-row", type=int, default=1, help="First row to scan. Default: 1.")
    parser.add_argument("--csv", default=None, help="CSV from parser. Default: result.csv, then results.csv.")
    parser.add_argument("--search-dir", default=".", help="Directory for Excel search. Default: current dir.")
    parser.add_argument("--output", default=None, help="Output .xlsx path. Default: *_updated.xlsx.")
    parser.add_argument(
        "--value-divisor",
        type=float,
        default=1.0,
        help="Divide CSV values before writing to Excel. Default: 1.0.",
    )
    return parser


def main():
    parser = build_arg_parser()
    args = parser.parse_args()

    csv_path = args.csv or find_default_csv()
    mapping = load_mapping_from_csv(csv_path)
    if not mapping:
        raise ValueError(f"No data loaded from {csv_path}")

    excel_file = args.excel_file
    if not excel_file:
        # Пытаемся найти любой подходящий Excel рядом, если путь не указан
        try:
            excel_file = resolve_excel_file(".xlsx", args.search_dir)
        except Exception:
            raise ValueError("No Excel file provided and none found in current directory.")

    excel_file = resolve_excel_file(excel_file, args.search_dir)
    print(f"Loaded {len(mapping)} salary values from {csv_path}")
    print(f"Excel file: {excel_file}")
    update_excel_with_mapping(
        file_path=excel_file,
        sheet_name=args.sheet,
        name_col=column_to_index(args.name_col),
        target_col=column_to_index(args.target_col),
        mapping=mapping,
        start_row=args.start_row,
        output_path=args.output,
        value_divisor=args.value_divisor,
    )


if __name__ == "__main__":
    main()
